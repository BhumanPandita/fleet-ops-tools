"""
MRO Labour Task Clustering Pipeline
------------------------------------
Goal: For Labour Cost rows, cluster semantically similar task descriptions
(written differently across MROs) into fine-grained task groups, then find
the minimum man hours per task cluster.
"""

import re
import warnings
import numpy as np
import pandas as pd
from sentence_transformers import SentenceTransformer
from sklearn.cluster import AgglomerativeClustering
from sklearn.preprocessing import normalize

warnings.filterwarnings("ignore")

# ── Configuration ─────────────────────────────────────────────────────────────
INPUT_FILE  = "Combined Sheet New Planes.xlsx"
OUTPUT_FILE = "MRO_Labour_Task_Clusters.xlsx"
SHEET_NAME  = "Combined"

# Cosine distance threshold for agglomerative clustering.
# Lower = more clusters (finer grain). Range 0–2; 0.35 is a good starting point.
DISTANCE_THRESHOLD = 0.35

# Sentence-transformer model (small, fast, aviation-friendly enough)
EMBEDDING_MODEL = "all-MiniLM-L6-v2"
# ──────────────────────────────────────────────────────────────────────────────


def clean_description(text: str) -> str:
    """Normalise free-text task descriptions before embedding."""
    if not isinstance(text, str):
        return ""
    text = text.upper().strip()
    # collapse whitespace and remove special chars that add no meaning
    text = re.sub(r"[^A-Z0-9\s/\-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def load_labour_rows(path: str) -> pd.DataFrame:
    print(f"Loading '{path}' …")
    df = pd.read_excel(path, sheet_name=SHEET_NAME)

    # Normalise column names (strip spaces and embedded newlines)
    df.columns = [re.sub(r'\s+', ' ', c).strip() for c in df.columns]

    print(f"  Total rows: {len(df):,}")

    # Filter Labour Cost only
    labour_mask = df["Material / Services"].astype(str).str.strip().str.lower() == "labor cost"
    df_labour = df[labour_mask].copy()
    print(f"  Labour Cost rows: {len(df_labour):,}")

    # Drop rows where Description is empty
    df_labour = df_labour[df_labour["Description"].notna()].copy()

    # Coerce Man Hour / Qty to numeric; non-numeric (e.g. 'FIXED COST') → NaN
    df_labour["Man Hour / Qty"] = pd.to_numeric(df_labour["Man Hour / Qty"], errors="coerce")

    # Drop rows with zero or missing man hours — no useful data for min-hours analysis
    df_labour = df_labour[df_labour["Man Hour / Qty"] > 0].copy()
    print(f"  After dropping zero/missing man hours: {len(df_labour):,}")

    df_labour["description_clean"] = df_labour["Description"].apply(clean_description)
    return df_labour.reset_index(drop=True)


def embed_unique_descriptions(descriptions: list[str]) -> np.ndarray:
    print(f"\nLoading embedding model '{EMBEDDING_MODEL}' …")
    model = SentenceTransformer(EMBEDDING_MODEL)
    print(f"Embedding {len(descriptions):,} unique descriptions …")
    embeddings = model.encode(
        descriptions,
        batch_size=256,
        show_progress_bar=True,
        convert_to_numpy=True,
    )
    # L2-normalise so cosine distance = euclidean distance on unit sphere
    return normalize(embeddings, norm="l2")


def cluster_descriptions(embeddings: np.ndarray) -> np.ndarray:
    print(f"\nClustering (distance_threshold={DISTANCE_THRESHOLD}) …")
    clusterer = AgglomerativeClustering(
        n_clusters=None,
        metric="cosine",
        linkage="average",          # average linkage works well for text
        distance_threshold=DISTANCE_THRESHOLD,
    )
    labels = clusterer.fit_predict(embeddings)
    n_clusters = len(set(labels))
    print(f"  Formed {n_clusters:,} clusters from {len(embeddings):,} unique descriptions")
    return labels


def build_cluster_label(group: pd.DataFrame) -> str:
    """
    Pick the most representative (shortest, clearest) description in a cluster
    as the canonical cluster label.
    """
    shortest = group.loc[group["description_clean"].str.len().idxmin(), "Description"]
    return str(shortest).strip()


def aggregate_results(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      summary  – one row per cluster: canonical label, count, min/mean/max man hours,
                 and the Tail + Card/WO of the minimum-man-hour row.
      detailed – original rows with cluster_id and cluster_label appended.
    """
    # ── Summary ──────────────────────────────────────────────────────────────
    summary_rows = []
    for cluster_id, grp in df.groupby("cluster_id"):
        min_idx   = grp["Man Hour / Qty"].idxmin()
        min_row   = grp.loc[min_idx]
        min_mh    = grp["Man Hour / Qty"].min()
        mro_min = (
            grp.groupby("MRO")["Man Hour / Qty"].min()
            .sort_values()
            .reset_index()
            .rename(columns={"MRO": "MRO Name", "Man Hour / Qty": "MRO Min MH"})
        )
        summary_rows.append({
            "Cluster ID":           cluster_id,
            "Cluster Label":        grp["cluster_label"].iloc[0],
            "Task Count":           len(grp),
            "Unique Descriptions":  grp["description_clean"].nunique(),
            "MROs Present":         ", ".join(sorted(grp["MRO"].dropna().unique())),
            "Min Man Hours":        min_mh,
            # Negotiation window: start low (Min − 40%) and settle no higher
            # than Min − 15%.
            "Negotiation Start (Min −40%)": round(min_mh * 0.60, 2),
            "Negotiation End (Min −15%)":   round(min_mh * 0.85, 2),
            "Mean Man Hours":       round(grp["Man Hour / Qty"].mean(), 2),
            "Max Man Hours":        grp["Man Hour / Qty"].max(),
            "Min MH — MRO":         min_row["MRO"],
            "Min MH — Tail":        min_row["Tail"],
            "Min MH — Card/WO":     min_row["Card / WO"],
            "Min MH — Sheet Ref":   min_row["Sheet ref"],
            "Min MH — Description": min_row["Description"],
            "MRO Breakdown (min MH each)": " | ".join(
                f"{r['MRO Name']}: {r['MRO Min MH']}" for _, r in mro_min.iterrows()
            ),
        })

    summary = (
        pd.DataFrame(summary_rows)
        .sort_values("Cluster Label")
        .reset_index(drop=True)
    )

    # ── Detailed ─────────────────────────────────────────────────────────────
    detailed = df.drop(columns=["description_clean"]).copy()
    detailed = detailed.sort_values(["cluster_id", "Man Hour / Qty"]).reset_index(drop=True)

    return summary, detailed


def save_output(summary: pd.DataFrame, detailed: pd.DataFrame, path: str) -> None:
    from excel_search import add_search_sheet
    from openpyxl.utils import get_column_letter

    # Columns to KEEP (in this order) on the "All Labour Rows" sheet — the rest
    # are still written but hidden, so the ground team see a lean view.
    ALR_KEEP = [
        "cluster_id",       # Cluster Id
        "Tail",
        "MRO",
        "cluster_label",    # Cluster Label
        "Description",
        "Man Hour / Qty",   # Man Hours
        "Amount $",         # Amount
        "Labor rate",       # Labor Rate
    ]
    # Columns to hide on the "Search Clusters" search sheet.
    SC_HIDE = [
        "Task Count",
        "Unique Descriptions",
        "Mean Man Hours",
        "Min MH — Card/WO",
        "Min MH — Sheet Ref",
    ]

    # Reorder detailed so the kept columns come first (kept order), rest after.
    ordered = [c for c in ALR_KEEP if c in detailed.columns]
    rest    = [c for c in detailed.columns if c not in ordered]
    detailed = detailed[ordered + rest]

    print(f"\nSaving results to '{path}' …")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Cluster Summary", index=False)
        detailed.to_excel(writer, sheet_name="All Labour Rows", index=False)

        # Auto-fit column widths for summary sheet
        ws = writer.sheets["Cluster Summary"]
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        # ── Order-independent keyword search sheets (for the ground team) ──────
        # Inserted at the front so the file opens on a search box.
        add_search_sheet(
            writer.book,
            title="Search Rows",
            src_sheet="All Labour Rows",
            headers=list(detailed.columns),
            search_col_name="Description",
            blurb="Searches every labour row — trace a match back via Tail + Card/WO.",
            n_data_rows=len(detailed),
        )
        add_search_sheet(
            writer.book,
            title="Search Clusters",
            src_sheet="Cluster Summary",
            headers=list(summary.columns),
            search_col_name="Cluster Label",
            tab_color="70AD47",
            blurb="Searches the cluster names to find a task group and its min man hours.",
            n_data_rows=len(summary),
        )

        # ── Visual tidy-up for the ground team ────────────────────────────────
        # Hide the sheets they don't need to open directly.
        for name in ("Search Rows", "Cluster Summary"):
            writer.book[name].sheet_state = "hidden"

        # Search Clusters: hide the non-essential summary columns.
        sc_ws   = writer.book["Search Clusters"]
        sum_cols = list(summary.columns)
        for name in SC_HIDE:
            if name in sum_cols:
                sc_ws.column_dimensions[
                    get_column_letter(sum_cols.index(name) + 1)
                ].hidden = True

        # All Labour Rows: hide every column after the kept block. Helper
        # columns (__match/__rank) are already hidden by add_search_sheet.
        alr_ws = writer.book["All Labour Rows"]
        for idx in range(len(ordered) + 1, len(detailed.columns) + 1):
            alr_ws.column_dimensions[get_column_letter(idx)].hidden = True

    print(f"Done. Output: {path}")


def main():
    import os
    os.chdir("/Users/bhumanpandita/Documents/IndiGo/MRO")

    # 1. Load & filter
    df = load_labour_rows(INPUT_FILE)

    # 2. Get unique cleaned descriptions for embedding (avoids embedding duplicates)
    unique_descs = df["description_clean"].unique().tolist()

    # 3. Embed
    embeddings = embed_unique_descriptions(unique_descs)

    # 4. Cluster unique descriptions
    labels = cluster_descriptions(embeddings)

    # 5. Map cluster labels back to every row
    desc_to_cluster = dict(zip(unique_descs, labels))
    df["cluster_id"] = df["description_clean"].map(desc_to_cluster)

    # 6. Build human-readable cluster label (shortest representative description)
    cluster_label_map = (
        df.groupby("cluster_id")
        .apply(build_cluster_label, include_groups=False)
        .to_dict()
    )
    df["cluster_label"] = df["cluster_id"].map(cluster_label_map)

    # 7. Aggregate
    summary, detailed = aggregate_results(df)

    print(f"\n{'─'*60}")
    print(f"  Labour rows analysed : {len(df):,}")
    print(f"  Unique descriptions  : {len(unique_descs):,}")
    print(f"  Task clusters formed : {summary['Cluster ID'].nunique():,}")
    print(f"{'─'*60}")
    print("\nTop 10 clusters by task count:")
    print(
        summary[["Cluster Label", "Task Count", "Min Man Hours", "Max Man Hours"]]
        .sort_values("Task Count", ascending=False)
        .head(10)
        .to_string(index=False)
    )

    # 8. Save
    save_output(summary, detailed, OUTPUT_FILE)


if __name__ == "__main__":
    main()
