"""
MRO Categorization Script
Categorizes each row in 'RAW DATA' sheet using:
1. Numeric Part No → exact lookup in Categorisations sheet
2. Alpha-numeric Part No → Part Description + Description keyword matching
3. '-' Part No (Labor/Supporting Cost) → Description keyword matching + ATA chapter name matching
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import re
from collections import defaultdict

INPUT_FILE = "Combined RAW DATA PLANES 10.xlsx"
OUTPUT_FILE = "Combined RAW DATA PLANES 10 - Categorized.xlsx"

# ─────────────────────────────────────────────
# 1. LOAD MAPPING TABLES FROM CATEGORISATIONS SHEET
# ─────────────────────────────────────────────

print("Loading categorisation mappings...")
wb_ref = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
ws_cat = wb_ref["Categorisations"]

part_no_map = {}   # str(numeric_id) -> (Category, SubCategory)
ata_map = {}       # ata_chapter_str -> (ata_name, Category, SubCategory)

for row in ws_cat.iter_rows(min_row=4, values_only=True):
    # Part No mapping (cols A-C, indices 0-2)
    pn_val, cat, subcat = row[0], row[1], row[2]
    if pn_val is not None and cat is not None:
        key = str(int(pn_val)) if isinstance(pn_val, (int, float)) else str(pn_val).strip()
        part_no_map[key] = (str(cat).strip(), str(subcat).strip() if subcat else "")

    # ATA chapter mapping (cols H-K, indices 7-10)
    ata_ch, ata_name, ata_cat, ata_subcat = row[7], row[8], row[9], row[10]
    if ata_ch is not None and ata_cat is not None:
        if isinstance(ata_ch, (int, float)):
            key = str(int(ata_ch))
        else:
            key = str(ata_ch).strip()
        ata_map[key] = (
            str(ata_name).strip().upper() if ata_name else "",
            str(ata_cat).strip(),
            str(ata_subcat).strip() if ata_subcat else ""
        )

# Build ATA name → (Category, SubCategory) for description matching
# (longest names first to avoid partial-match shadowing)
ata_name_to_cat = {}
for key, (ata_name, cat, subcat) in ata_map.items():
    if ata_name and cat and cat not in ("Non MPD Task", "GENERAL"):
        ata_name_to_cat[ata_name] = (cat, subcat)

ata_names_sorted = sorted(ata_name_to_cat.keys(), key=len, reverse=True)

print(f"  Part No mappings loaded: {len(part_no_map)}")
print(f"  ATA Chapter mappings loaded: {len(ata_map)}")
wb_ref.close()

# ─────────────────────────────────────────────
# 2. KEYWORD RULES
# ─────────────────────────────────────────────

# Each entry: (regex_pattern, Category, SubCategory)
# Applied in ORDER – first match wins.

PART_DESC_RULES = [
    # POWERPLANT – APU
    (r"\bAPU\b", "POWERPLANT", "APU"),
    # POWERPLANT – Thermal Blanket
    (r"THERMAL BLANKET|INSULATION BLANKET", "POWERPLANT", "Thermal Blanket"),
    # POWERPLANT – Nacelle Spares
    (r"\bNACELLE\b|\bCOWL\b|\bINLET COWL\b|\bTHRUST REVERSER\b|\bCASCADE\b|\bFAN COWL\b|\bBLOCKER DOOR\b", "POWERPLANT", "Nacelle Spares"),
    # POWERPLANT – Engine
    (r"\bENGINE\b|\bTURBINE\b|\bCFM\b|\bFAN BLADE\b|\bCOMBUSTOR\b|\bNOZZLE\b|\bVANE\b|\bSHROUD\b|\bSEAL PLATE\b|EXHAUST|\bOIL PUMP\b|\bFUEL PUMP\b|\bGEARBOX\b|\bMUFFLE\b", "POWERPLANT", "Engine"),

    # LANDING GEAR
    (r"\bLANDING GEAR\b|\bNLG\b|\bMLG\b|\bWHEEL\b|\bBRAKE\b|\bTIRE\b|\bTYRE\b|\bAXLE\b|\bSHIMMY\b|\bTORQUE LINK\b|\bDRAG BRACE\b|\bSIDESTAY\b|\bOLEO\b|\bSHOCK STRUT\b", "LANDING GEAR", "Landing Gear"),

    # WINGS & PYLONS
    (r"\bSLAT\b|\bFLAP\b|\bSPOILER\b|\bAILERON\b|\bPYLON\b|\bWING TIP\b|\bLEADING EDGE\b|\bTRAILING EDGE\b|\bWINGLET\b", "WINGS & PYLONS", "Wing/Pylon Spares"),
    (r"\bWING\b", "WINGS & PYLONS", "Wing/Pylon Spares"),

    # CABIN – Lavatory
    (r"\bLAVATORY\b|\bLAV\b|\bTOILET\b|\bWASTE\b|\bWATER WASTE\b", "CABIN", "Lavatory"),
    # CABIN – Galley
    (r"\bGALLEY\b|\bTROLLEY\b|\bGALLEY INSERT\b", "CABIN", "Galley"),
    # CABIN – Pax Seats
    (r"\bSEAT PAN\b|\bSEAT COVER\b|\bARMREST\b|\bARMCAP\b|\bBACKREST\b|\bHEADREST\b|\bCUSHION\b|\bTRAY TABLE\b|\bSEAT BELT\b|\bSEAT RAIL\b|\bRECLINE\b|\bPASSENGER SEAT\b|\bPAX SEAT\b", "CABIN", "Pax Seats"),
    # CABIN – Pax Compartment
    (r"\bLIFE VEST\b|\bOXYGEN MASK\b|\bOXYGEN BOTTLE\b|\bCARPET\b|\bCABIN PANEL\b|\bOVERHEAD BIN\b|\bPSU\b|\bREADING LIGHT\b|\bPASSENGER SERVICE\b|SLIDE.*RAFT|SLIDE AFT|SLIDE FW|ESCAPE SLIDE|HANDSET", "CABIN", "Pax Compartment"),

    # CARGO
    (r"\bCARGO\b|\bCONTAINER\b|\bPALLET\b|\bULD\b|\bCARGO LINER\b", "CARGO", "Cargo Spares"),

    # COCKPIT
    (r"\bCOCKPIT\b|\bINSTRUMENT PANEL\b|\bGLARESHIELD\b|\bCONTROL PANEL\b|\bOVERHEAD PANEL\b", "COCKPIT", "Cockpit Furnishings"),

    # PAINTS (before Consumables to avoid THINNER going to wrong bucket)
    (r"\bPAINT\b|BAC70[0-9]|EXTERIOR FINISH|ZINC CHROMATE|CLEARCOAT|CLEAR COAT", "PAINTS", "Exterior Paints"),
    (r"\bPRIMER\b|\bALODINE\b|\bCHROMATION\b|CONVERSION COAT|EPOXY COAT|\bCHROMIC ACID\b", "PAINTS", "Exterior Paints"),
    (r"\bCOATING CONVERSION\b|\bCOATING\b|\bVARNISH\b|\bLACQUER\b|\bPOLYURETHANE\b|FREEKOTE|INTERIOR PAINT|INTERIOR COAT", "PAINTS", "Exterior Paints"),
    (r"THINNER|THINNERS?\s*C\d|THINNERC\d", "PAINTS", "Exterior Paints"),
    (r"\bPUTTY\b", "PAINTS", "Exterior Paints"),

    # GENERAL – Hardwares (fasteners and fittings – checked BEFORE consumables)
    (r"\bRIVET\b|\bSCREW\b|\bBOLT\b|\bNUT\b|\bWASHER\b|\bCOLLAR\b|\bSTUD\b", "GENERAL", "Hardwares"),
    (r"HI[-\s]LOK|HILOK|\bNUTPLATE\b|\bHEX BOLT\b|\bFASTENER\b|\bLOCKNUT\b|\bSPACER\b|\bSLEEVE\b|\bFERRULE\b", "GENERAL", "Hardwares"),
    (r"COTTER[-\s]?PIN|PIN[-,\s]COTTER|PIN,\s*COTTER|\bCOTTER\b", "GENERAL", "Hardwares"),
    (r"LOCK[-\s]?WIRE|WIRE[-\s]LOCK|SAFETY[-\s]WIRE|WIRE LOCKING|LOCKING WIRE", "GENERAL", "Hardwares"),
    (r"\bINSERT\b|\bBUSHING\b|\bBEARING\b|\bCLAMP\b|\bFITTING\b|\bCAP SCREW\b|\bEYEBOLT\b", "GENERAL", "Hardwares"),
    (r"\bRING\b|\bCAP\b|\bPLUG\b|\bPIN\b", "GENERAL", "Hardwares"),

    # GENERAL – Consumables (chemicals, materials, supplies)
    (r"\bSEALANT\b|\bSEAL\b|\bPACKING\b|\bO-RING\b|\bGASKET\b|\bRUBBER\b|\bGROMET\b|\bFILLET\b", "GENERAL", "Consumables"),
    (r"\bSOLVENT\b|\bIPA\b|\bMEK\b|\bACETONE\b|ISOPROPYL|METHYL ETHYL|CLEANING FLUID|CLEANING CLOTH|CLEAN WIPE|LOTOXANE|CLEANER\b|EXT\.\s*CLEANER", "GENERAL", "Consumables"),
    (r"\bADHESIVE\b|\bTAPE\b|\bBOSTIK\b|\bSILICONE\b|\bRTV\b|BAGGING FILM|BAGG\b|\bCOMPOUND\b|\bGREASE\b|\bLUBRICANT\b|\bHYDRAULIC FLUID\b|HYJET|HYD FLUID", "GENERAL", "Consumables"),
    (r"\bACTIVATOR\b|HARDENER|ACTIVATOR\b|\bWATER\b|\bCLOTH\b|\bCOTTON\b|\bFOAM\b|NITRIL|NITRILE|GLOVE|\bFILTER\b|\bWIPE\b|\bNAPKIN\b", "GENERAL", "Consumables"),
    (r"\bNYLON\b|NAYLON|TORBA\b|BEZI\b|FIRCA\b|TULBENT|BOSTIK|KARTI\b|GECICI|DIKKAT|VAZELINE|VASELINE|\bOIL\b", "GENERAL", "Consumables"),
    (r"AIRWEAVE|BLEEDER|RELEASE AGENT|PTFE|SCOTCH.BRITE|ABRASIVE|FILLER.*CORE|CORE.*FILLER|MILLING CUTTER|MINERAL OIL", "GENERAL", "Consumables"),
    (r"LEAD.BONDING|BONDING AGENT|\bFILM\b|\bTORBA\b|BEZI|ŞEFFAF|KARBONMASKE|\bMASKE\b|ELDIVEN|FIRÇA|FIRCASI", "GENERAL", "Consumables"),
    (r"PLACARD|DECAL|MARKING|LABEL|STICKER|SIGN\b", "GENERAL", "Consumables"),

    # AIRFRAME – Structures (structural parts)
    (r"\bPLATE THS\b|\bTHS PLATE\b|\bPLATE\b|\bPANEL\b|\bSKIN\b|\bBRACKET\b|\bSTRAP\b|\bPATCH\b|\bDOUBLER\b|\bSTRINGER\b|\bFRAME\b|\bBULKHEAD\b|\bANGLE\b|\bSHIM\b|\bRIB\b|\bSPAR\b|\bFLOOR BEAM\b|\bFLOOR PANEL\b|\bSECTION\b|\bBELLOWS\b|\bFAIRING\b|\bCOVER\b", "AIRFRAME", "Structures"),

    # COMPONENTS (avionics / systems LRUs)
    (r"\bVALVE\b|SAFETY VAL\b|\bPUMP\b|\bACTUATOR\b|\bSENSOR\b|\bSWITCH\b|\bREGULATOR\b|\bINDICATOR\b|\bTRANSMITTER\b|\bTRANSDUCER\b|\bDETECTOR\b|\bCOMPUTER\b|\bRELAY\b|\bCONTROLLER\b|\bCONTROL BOX\b|\bLRU\b|\bECU\b|\bFMGC\b|\bADIRU\b|\bELECTRIC\b|\bELECTRONIC\b|\bAVIONIC\b", "COMPONENTS", "Component"),
    (r"\bHARNESS\b|\bWIRE BUNDLE\b|\bCONNECTOR\b|\bCONDUIT\b|\bCOMPOSITE\b|\bTRANSFORMER\b|\bGENERATOR\b|\bMOTOR\b", "COMPONENTS", "Component"),
    (r"HEAT EXCHANGER|CONDENSER|REHEATER|CSAS|ANTICOLLISION|ANTI.COLLISION|BEACON|BATTERY|FLIGHT DATA RECORDER|FDR\b|CVR\b|RECORDER|YAW DAMPER|SPRING ROD|STRUT|HANDSET", "COMPONENTS", "Component"),

    # PRESERVATION
    (r"PRESERVATION|DEPRESERVATION", "PRESERVATION/ DEPRESERVATION", "Airframe"),
]

# Compile patterns
PART_DESC_RULES_COMPILED = [
    (re.compile(pattern, re.IGNORECASE), cat, subcat)
    for pattern, cat, subcat in PART_DESC_RULES
]

# ─── Description-level keyword rules ──────────────────────────────────────────

DESC_RULES = [
    # Preservation/storage
    (r"PRESERVATION|DEPRESERV|STORAGE.*CHECK|PARKING.*CHECK|PERIODIC.*GROUND.*CHECK|RETURN TO SERVICE", "PRESERVATION/ DEPRESERVATION", "Airframe"),
    (r"APU.*PRESERVATION|PRESERVATION.*APU", "PRESERVATION/ DEPRESERVATION", "APU"),
    (r"ENGINE.*PRESERVATION|PRESERVATION.*ENGINE", "PRESERVATION/ DEPRESERVATION", "Engine"),

    # Paints
    (r"PAINT STRIP|REPAINT|RE-PAINT|BARE METAL INSP|BMI\b|EXTERIOR PAINT|STRIP.*PAINT|PAINT.*FUSELAGE", "AIRFRAME", "BMI"),
    (r"\bPAINT\b|\bPAINTING\b", "PAINTS", "Exterior Paints"),

    # Cargo
    (r"CARGO COMPARTMENT|BULK CARGO|CARGO LINER|AFT CARGO|FWD CARGO|CARGO DOOR|CARGO PANEL", "CARGO", "Cargo Spares"),

    # Landing gear
    (r"LANDING GEAR|NOSE LANDING|MAIN LANDING|NLG\b|MLG\b|\bWHEEL\b|\bBRAKE\b|\bTIRE\b|\bTYRE\b|SHOCK ABSORBER|SHIMMY|TORQUE LINK|OLEO", "LANDING GEAR", "Landing Gear"),

    # Powerplant – APU
    (r"\bAPU\b", "POWERPLANT", "APU"),
    # Powerplant – Engine
    (r"\bENGINE\b|\bCFM\b|\bTURBINE\b|\bFAN BLADE\b|\bEXHAUST\b|\bNACELLE\b|\bCOWL\b|\bTHRUST REVERSER\b|OIL SHEET|OIL FINDING|OIL ANALYSIS|BORESCOPE", "POWERPLANT", "Engine"),

    # Wings & Pylons
    (r"\bSLAT\b|\bFLAP\b|\bSPOILER\b|\bAILERON\b|\bPYLON\b|\bWINGLET\b|\bLEADING EDGE\b|\bTRAILING EDGE\b|WING SKIN|WING RIB|WING SPAR|OUTER WING|INNER WING|I/B FLAP|O/B FLAP|WFX\b", "WINGS & PYLONS", "Repair"),

    # AIRFRAME – Structures / Corrosion
    (r"CORROSION|CORRODE", "AIRFRAME", "Corrosion"),
    (r"FUSELAGE|SKIN PANEL|LOWER SKIN|UPPER SKIN|SECTION 11|SECTION 12|SECTION 13|SECTION 14|SECTION 15|SECTION 16|SECTION 17|SECTION 18|SECTION 19|FR\d+|STGR\d+|STRINGER|FRAME REPAIR|DOUBLER|STRUCTURE REPAIR|SRM|STRUCTURAL|NICKS|DENT AND|SCRATCH|RUB MARK.*REPAIR", "AIRFRAME", "Structures"),

    # Cabin – Galley
    (r"\bGALLEY\b|\bGALLEY\s+\d+\b|G[1-9]\b", "CABIN", "Galley"),
    # Cabin – Lavatory
    (r"\bLAVATORY\b|\bLAV\b|\bTOILET\b|WATER WASTE|WASTE DISPOSAL", "CABIN", "Lavatory"),
    # Cabin – Seats
    (r"\bPAX SEAT\b|\bPASSENGER SEAT\b|REFURBISHMENT.*SEAT|SEAT REFURB|SEAT CUSHION|ARMREST|SEAT PAN", "CABIN", "Pax Seats"),
    # Cabin – General
    (r"CABIN FLOOR|CABIN STRUCTURE|CABIN PANEL|OVERHEAD BIN|INTERIOR|CABIN LINER|CARGO COMPARTMENT LINER", "CABIN", "Pax Compartment"),

    # Cockpit
    (r"\bCOCKPIT\b", "COCKPIT", "Cockpit Furnishings"),

    # Supporting services – HOTAC
    (r"HOTEL|ACCOMMODATION|RAMANI|HOSTELLING|LODGE|\bHOTAC\b", "SUPPORTING SERVICES", "HOTAC"),
    # Supporting services – Hangar/Apron
    (r"FOLLOW ME|CRANE|PARKING CHARGE|AIRCRAFT PARKING|GROUND SUPPORT|FUEL.*(SERVICE|OIL)|HYDRAULIC SERVICE|TRANSPORT|HANGAR|APRON|GSD SUPPORT|CUSTOM CLEARANCE|AIRCRAFT HANDLING|AIRCRAFT TOW|AIRCRAFT PUSH|MANPOWER SUPPORT|FUEL/OIL|FUEL\s*/\s*OIL", "SUPPORTING SERVICES", "Hangar/ Apron"),

    # Logistics
    (r"BOX FABRICATION|TRANSPORT|FREIGHT|COURIER|PACKAGING|SHIPPING|DELIVERY|LOGISTICS", "LOGISTICS & OUTSOURCE", "Logistics"),

    # Outsource
    (r"OUTSOURCE|SUBCONTRACT|EXTERNAL VENDOR|THIRD PARTY", "LOGISTICS & OUTSOURCE", "Outsource"),

    # Tech services
    (r"INSPECTION|TECH SERVICE|ENGINEERING ORDER|EO\b|SB\b|SERVICE BULLETIN|AIRWORTHINESS DIRECTIVE|AD\b|MODIFICATION|MOD\b|DOCUMENTATION|CALIBRATION|NDT\b|BORESCOPE.*INSP|ENGINEERING WORK", "TECH SERVICES", "Tech Services"),
]

DESC_RULES_COMPILED = [
    (re.compile(pattern, re.IGNORECASE), cat, subcat)
    for pattern, cat, subcat in DESC_RULES
]

# ─────────────────────────────────────────────
# 3. CATEGORIZATION FUNCTION
# ─────────────────────────────────────────────

def match_ata_name(text):
    """Try to match text against ATA chapter names."""
    if not text:
        return None
    text_up = text.upper()
    for ata_name in ata_names_sorted:
        if ata_name and ata_name in text_up:
            return ata_name_to_cat[ata_name]
    return None


def apply_keyword_rules(text, rules):
    """Apply compiled regex rules to text. Returns (Category, SubCategory) or None."""
    if not text:
        return None
    for pattern, cat, subcat in rules:
        if pattern.search(text):
            return (cat, subcat)
    return None


def categorize_row(mat_svc, desc, part_no, part_desc):
    """
    Returns (Category, SubCategory, method_used).
    method_used is a string for traceability.
    """
    mat_svc = str(mat_svc or "").strip()
    desc = str(desc or "").strip()
    part_no = str(part_no or "").strip()
    part_desc = str(part_desc or "").strip()

    # ── Step 1: Part No exact lookup (direct string + numeric normalization) ────
    _SKIP = {"uncategorised", "uncategorized", "non traceable", "credit"}
    if part_no and part_no != "-":
        pn_str = part_no.strip()

        # Direct string match — handles alpha-numeric PNs stored in the map
        if pn_str in part_no_map:
            cat, subcat = part_no_map[pn_str]
            if cat and cat.lower() not in _SKIP:
                return (cat, subcat, "PartNo-lookup")

        # Numeric normalization — handles int/float format mismatches
        try:
            num_val = float(pn_str)
            if num_val != float("inf") and num_val == int(num_val):
                keys_to_try = [str(int(num_val)), pn_str, pn_str.lstrip("0") or "0"]
                for key in keys_to_try:
                    if key in part_no_map:
                        cat, subcat = part_no_map[key]
                        if cat and cat.lower() not in _SKIP:
                            return (cat, subcat, "PartNo-lookup")
        except (ValueError, OverflowError):
            pass

    # ── Step 2: rows where Part No is '-' (Labor / Supporting / general) ──────
    if part_no == "-" or not part_no:
        # 2a. Supporting Cost with Part No = '-'
        if mat_svc == "Supporting Cost":
            result = apply_keyword_rules(desc, DESC_RULES_COMPILED)
            if result:
                return (result[0], result[1], "SupportCost-DescKW")
            result = apply_keyword_rules(desc, PART_DESC_RULES_COMPILED)
            if result:
                return (result[0], result[1], "SupportCost-PartDescKW")
            # Default for supporting cost
            return ("SUPPORTING SERVICES", "Hangar/ Apron", "SupportCost-default")

        # 2b. Labor Cost / Labor+Material Cost
        if mat_svc in ("Labor Cost", "Labor + Material Cost"):
            # Skip spreadsheet formula references
            if desc.startswith("='"):
                return ("TECH SERVICES", "Tech Services", "LaborCost-formula-default")

            # Try ATA name matching first
            result = match_ata_name(desc)
            if result:
                return (result[0], result[1], "LaborCost-ATA")

            # Try description keyword rules
            result = apply_keyword_rules(desc, DESC_RULES_COMPILED)
            if result:
                return (result[0], result[1], "LaborCost-DescKW")

            return ("TECH SERVICES", "Tech Services", "LaborCost-default")

        # 2c. Material Cost with Part No = '-' – categorize by description
        if mat_svc == "Material Cost":
            result = match_ata_name(desc)
            if result:
                return (result[0], result[1], "MatCost-dash-ATA")
            result = apply_keyword_rules(desc, DESC_RULES_COMPILED)
            if result:
                return (result[0], result[1], "MatCost-dash-DescKW")
            result = apply_keyword_rules(desc, PART_DESC_RULES_COMPILED)
            if result:
                return (result[0], result[1], "MatCost-dash-PartDescKW")
            return ("GENERAL", "Consumables", "MatCost-dash-default")

    # ── Step 3: alpha-numeric Part No ─────────────────────────────────────────
    # Primary signal: Part Description
    if part_desc and part_desc != "-":
        result = apply_keyword_rules(part_desc, PART_DESC_RULES_COMPILED)
        if result:
            return (result[0], result[1], "PartDesc-KW")

    # Secondary signal: Description
    if desc and desc != "-" and not desc.startswith("='"):
        # Try ATA name
        result = match_ata_name(desc)
        if result:
            return (result[0], result[1], "AlphaPN-DescATA")
        # Try desc keywords
        result = apply_keyword_rules(desc, DESC_RULES_COMPILED)
        if result:
            return (result[0], result[1], "AlphaPN-DescKW")
        # Try part desc keywords on description text
        result = apply_keyword_rules(desc, PART_DESC_RULES_COMPILED)
        if result:
            return (result[0], result[1], "AlphaPN-DescPDKW")

    # If numeric lookup failed earlier, try description for those rows too
    result = apply_keyword_rules(desc, DESC_RULES_COMPILED)
    if result:
        return (result[0], result[1], "Fallback-DescKW")

    # Last resort: categorize by material/service type
    if mat_svc == "Material Cost":
        if desc and desc != "-" and not desc.startswith("='"):
            result = match_ata_name(desc)
            if result:
                return (result[0], result[1], "MatCost-DescATA-fallback")
        if part_no and part_no != "-":
            return ("GENERAL", "Consumables", "MatCost-default")

    if mat_svc == "Supporting Cost":
        # SV-ENGINEERING and other non-standard supporting cost part numbers
        if "ENGINEER" in (part_no + " " + desc).upper():
            return ("TECH SERVICES", "Tech Services", "SupportCost-engr")
        return ("SUPPORTING SERVICES", "Hangar/ Apron", "SupportCost-pn-default")

    if mat_svc in ("Labor Cost", "Labor + Material Cost"):
        return ("TECH SERVICES", "Tech Services", "Labor-final-default")

    return ("UNCATEGORIZED", "Uncategorized", "no-match")


# ─────────────────────────────────────────────
# 4. PROCESS ALL ROWS
# ─────────────────────────────────────────────

print("Loading RAW DATA...")
wb_in = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
ws_raw = wb_in["RAW DATA"]

print("Reading all rows (this may take a moment)...")
all_rows = list(ws_raw.iter_rows(min_row=1, values_only=True))
wb_in.close()

header = list(all_rows[0])
data_rows = all_rows[1:]
print(f"  Total data rows: {len(data_rows)}")

# Process
results = []
method_counts = {}
for row in data_rows:
    if not any(row):
        results.append(("", "", "empty"))
        continue
    mat_svc = row[9]
    desc = row[8]
    part_no = row[11]
    part_desc = row[12]
    cat, subcat, method = categorize_row(mat_svc, desc, part_no, part_desc)
    # Normalize subcategory casing from ATA map (TECH SERVICES → Tech Services)
    if subcat == "TECH SERVICES":
        subcat = "Tech Services"
    results.append((cat, subcat, method))
    method_counts[method] = method_counts.get(method, 0) + 1

# ─────────────────────────────────────────────
# 5. WRITE OUTPUT FILE
# ─────────────────────────────────────────────

print("Writing output file...")
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "RAW DATA"

# Write header
ws_out.append(header)

# Styles
uncategorized_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
cat_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green

for i, (orig_row, (cat, subcat, method)) in enumerate(zip(data_rows, results), 2):
    row_out = list(orig_row)
    row_out[13] = cat    # Category (col 14)
    row_out[14] = subcat  # Sub Category (col 15)
    ws_out.append(row_out)
    if cat == "UNCATEGORIZED":
        for col in range(1, 16):
            ws_out.cell(row=i, column=col).fill = uncategorized_fill

# Also copy Categorisations sheet
wb_ref2 = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
ws_cat_src = wb_ref2["Categorisations"]
ws_cat_out = wb_out.create_sheet("Categorisations")
for row in ws_cat_src.iter_rows(values_only=True):
    ws_cat_out.append(list(row))
wb_ref2.close()

wb_out.save(OUTPUT_FILE)
print(f"Saved: {OUTPUT_FILE}")

# ─────────────────────────────────────────────
# 6. STATS REPORT
# ─────────────────────────────────────────────

from collections import Counter
cat_counter = Counter()
for cat, subcat, method in results:
    if cat:
        cat_counter[f"{cat} / {subcat}"] += 1

print("\n=== Categorization Method Breakdown ===")
for method, count in sorted(method_counts.items(), key=lambda x: -x[1]):
    print(f"  {count:6d}  {method}")

print("\n=== Category / SubCategory Distribution ===")
for cat_sub, count in sorted(cat_counter.items(), key=lambda x: -x[1]):
    print(f"  {count:6d}  {cat_sub}")

uncategorized = sum(1 for cat, _, _ in results if cat == "UNCATEGORIZED")
print(f"\nTotal rows: {len(results)}")
print(f"Uncategorized: {uncategorized} ({100*uncategorized/max(len(results),1):.1f}%)")
