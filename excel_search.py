"""
excel_search.py — add an order-independent keyword search sheet to a workbook.
-----------------------------------------------------------------------------
Built for the ground team who work straight off the Excel file (no laptop /
localhost / dashboard). They type keywords in ANY order into one yellow cell
and the matching rows appear below — live, as they type.

So searching "cockpit glass broken" finds "GLASS BROKEN IN COCKPIT".

WHY NOT FILTER / dynamic arrays?
--------------------------------
The obvious solution is one FILTER() formula that spills the matching rows.
But openpyxl cannot mark a formula as a true "dynamic array" (that needs hidden
OOXML metadata it does not write). When Excel opens such a file it adds implicit
intersection "@" operators (=@IF, @SEARCH, @range…) which collapse every array
to a single cell — the search returns only one row or #VALUE!.

So instead we use the rock-solid classic approach that needs NO dynamic arrays:

  1. Two hidden helper columns on the DATA sheet (one formula per row):
       __match  : 1 if this row matches ALL typed keywords, else 0
       __rank   : running count among matches (1,2,3,…), blank if no match
     __match is written as a legacy array (CSE) formula so SUMPRODUCT/TEXTSPLIT
     evaluate in array context WITHOUT Excel adding "@".

  2. The Search sheet has a fixed block of result rows, each a plain
     INDEX/MATCH formula that pulls the i-th matching row by its __rank.
     Plain per-row formulas — Excel never breaks these with "@".

Match logic (order-independent, case-insensitive):
  SUMPRODUCT(--ISNUMBER(SEARCH( TEXTSPLIT(keywords) , description )))
      == COLUMNS(TEXTSPLIT(keywords))
i.e. every typed keyword is found somewhere in the description.

NOTE on the _xlfn. prefix: openpyxl stores formulas verbatim, but the file
format requires newer functions (TEXTSPLIT) to carry this internal prefix or
Excel shows #NAME?. Excel displays the clean name to the user.
"""

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INPUT_CELL   = "B3"        # the yellow "type here" cell
HEADER_ROW   = 6           # results header row on the search sheet
DATA_ROW     = 7           # results start spilling (via INDEX/MATCH) here
RESULT_SLOTS = 1000        # max matching rows shown (refine search if more)

MATCH_HDR = "__match"      # hidden helper column headers on the data sheet
RANK_HDR  = "__rank"

_THIN = Side(style="thin", color="BBBBBB")


def strip_helper_columns(ws) -> None:
    """Remove any previously-added __match / __rank helper columns (idempotent)."""
    # Delete from the right so indices stay valid.
    for col in range(ws.max_column, 0, -1):
        if ws.cell(row=1, column=col).value in (MATCH_HDR, RANK_HDR):
            ws.delete_cols(col, 1)


def add_search_sheet(
    book,
    title: str,
    src_sheet: str,
    headers: list[str],
    search_col_name: str,
    n_data_rows: int,
    tab_color: str = "FFC000",
    blurb: str = "",
) -> None:
    """Insert a live keyword-search sheet (at the front) for `src_sheet`.

    book            : openpyxl Workbook
    title           : name for the new search sheet
    src_sheet       : name of the data sheet to search
    headers         : column headers of the data sheet (in order)
    search_col_name : which column to match keywords against
    n_data_rows     : number of data rows (excluding header) in src_sheet
    """
    ws_data    = book[src_sheet]
    strip_helper_columns(ws_data)                          # clean re-runs

    ncols      = len(headers)
    last_data  = n_data_rows + 1                           # data occupies rows 2..last_data
    search_idx = headers.index(search_col_name) + 1
    desc_col   = get_column_letter(search_idx)             # description column on data sheet

    # ── 1. Hidden helper columns on the data sheet ────────────────────────────
    flag_idx   = ncols + 1
    rank_idx   = ncols + 2
    flag_col   = get_column_letter(flag_idx)
    rank_col   = get_column_letter(rank_idx)
    ws_data.cell(row=1, column=flag_idx, value=MATCH_HDR)
    ws_data.cell(row=1, column=rank_idx, value=RANK_HDR)

    box = f"'{title}'!${INPUT_CELL[0]}${INPUT_CELL[1:]}"   # e.g. 'Search Rows'!$B$3
    kw  = f'_xlfn.TEXTSPLIT(TRIM({box})," ")'              # the typed keywords as an array

    for r in range(2, last_data + 1):
        # __match : 1 when EVERY keyword is found in this row's description, else 0.
        # Written as a legacy array (CSE) formula → array context, no "@".
        match_f = (
            f'=IF(TRIM({box})="",0,'
            f'IF(SUMPRODUCT(--ISNUMBER(SEARCH({kw},{desc_col}{r})))=COLUMNS({kw}),1,0))'
        )
        ws_data.cell(row=r, column=flag_idx,
                     value=ArrayFormula(f"{flag_col}{r}", match_f))
        # __rank : running position among matches (1,2,3,…), blank otherwise.
        ws_data.cell(row=r, column=rank_idx,
                     value=f'=IF({flag_col}{r}=1,SUM(${flag_col}$2:{flag_col}{r}),"")')

    ws_data.column_dimensions[flag_col].hidden = True
    ws_data.column_dimensions[rank_col].hidden = True

    # ── 2. The search sheet ───────────────────────────────────────────────────
    ws = book.create_sheet(title, 0)                       # 0 = put it first
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    last_col = get_column_letter(ncols)

    # Title + instructions
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = "🔍  TASK SEARCH  —  type keywords in ANY order, then press Enter"
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F4E78")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    sub = ws["A2"]
    sub.value = (
        "Example: typing  cockpit glass broken  also finds  “GLASS BROKEN IN "
        "COCKPIT”.   " + blurb
    ).strip()
    sub.font = Font(italic=True, size=10, color="555555")
    sub.alignment = Alignment(horizontal="left", indent=1)

    # Input label + yellow search box
    lbl = ws["A3"]
    lbl.value = "Search:"
    lbl.font = Font(bold=True, size=12)
    lbl.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("B3:F3")
    inp = ws[INPUT_CELL]
    inp.fill = PatternFill("solid", fgColor="FFF2CC")
    inp.font = Font(bold=True, size=12)
    inp.alignment = Alignment(vertical="center", indent=1)
    medium = Side(style="medium", color="BF9000")
    for col in range(2, 7):                                # B..F border box
        ws.cell(row=3, column=col).border = Border(
            left=medium, right=medium, top=medium, bottom=medium)
    ws.row_dimensions[3].height = 24

    # Live match counter
    box_ref = f"${INPUT_CELL[0]}${INPUT_CELL[1:]}"
    cnt = ws["A4"]
    cnt.value = (
        f'=IF(TRIM({box_ref})="","",'
        f'SUM(\'{src_sheet}\'!${flag_col}$2:${flag_col}${last_data})&'
        f'" match(es) found  (showing up to {RESULT_SLOTS})")'
    )
    cnt.font = Font(bold=True, size=11, color="C00000")
    ws.merge_cells(f"A4:{last_col}4")

    # Results header row (copied from the data sheet)
    for j, name in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=j, value=str(name))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2F5496")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    # Result rows: each cell pulls the i-th matching row via INDEX/MATCH.
    src   = f"'{src_sheet}'"
    slots = min(RESULT_SLOTS, n_data_rows)
    for i in range(slots):
        out_row = DATA_ROW + i
        nth     = f"ROWS(A${DATA_ROW}:A{out_row})"         # 1,2,3,… as it fills down
        for j in range(1, ncols + 1):
            col = get_column_letter(j)
            f = (
                f'=IFERROR(INDEX({src}!{col}$2:{col}${last_data},'
                f'MATCH({nth},{src}!${rank_col}$2:${rank_col}${last_data},0)),"")'
            )
            ws.cell(row=out_row, column=j, value=f)

    # Cosmetics
    for j in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.freeze_panes = f"A{DATA_ROW}"
