"""
add_search.py — add the order-independent search tabs to an EXISTING workbook
without re-running the clustering / embeddings.

Usage:
    python3 add_search.py                              # patches MRO_Labour_Task_Clusters.xlsx in place
    python3 add_search.py "Some Other File.xlsx"       # patches a different file in place

Expects the workbook to contain the sheets:
    'All Labour Rows'  (with a 'Description' column)
    'Cluster Summary'  (with a 'Cluster Label' column)
"""

import sys
import openpyxl
from excel_search import add_search_sheet

DEFAULT_FILE = "MRO_Labour_Task_Clusters.xlsx"


def _headers(ws):
    return [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]


def _data_rows(ws):
    # Scan from the bottom to find the last row that has any data,
    # ignoring blank cells anywhere in the middle of the sheet.
    last = ws.max_row
    while last > 1:
        if any(c.value is not None for c in ws[last]):
            break
        last -= 1
    return last - 1  # subtract the header row


def main(path: str) -> None:
    print(f"Opening '{path}' …")
    wb = openpyxl.load_workbook(path)

    # Remove old search tabs if this is a re-run, then rebuild them.
    for title in ("Search Rows", "Search Clusters"):
        if title in wb.sheetnames:
            del wb[title]

    rows_ws = wb["All Labour Rows"]
    clus_ws = wb["Cluster Summary"]

    add_search_sheet(
        wb, "Search Rows", "All Labour Rows",
        _headers(rows_ws), "Description", _data_rows(rows_ws),
        blurb="Searches every labour row — trace a match back via Tail + Card/WO.",
    )
    add_search_sheet(
        wb, "Search Clusters", "Cluster Summary",
        _headers(clus_ws), "Cluster Label", _data_rows(clus_ws),
        tab_color="70AD47",
        blurb="Searches the cluster names to find a task group and its min man hours.",
    )

    wb.save(path)
    print(f"Done. Search tabs added. Sheet order: {wb.sheetnames}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE)
